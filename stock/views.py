from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, DetailView, TemplateView, CreateView
from django.urls import reverse_lazy
from bootstrap_modal_forms.generic import BSModalCreateView, BSModalUpdateView

from rest_framework import viewsets
import openpyxl as px
import datetime

from .models import *
from device.models import *
from user.models import *
from .serializers import *
from .forms import *
from .filters import *


class OptionViewSet(viewsets.ModelViewSet):
	queryset = Option.objects.all()
	serializer_class = OptionSerializer
	filter_class = OptionFilter


class StorageItemViewSet(viewsets.ModelViewSet):
	queryset = StorageItem.objects.all()
	serializer_class = StorageItemSerializer
	filter_class = StorageItemFilter


class StorageCartViewSet(viewsets.ModelViewSet):
	queryset = StorageCart.objects.all()
	serializer_class = StorageCartSerializer
	filter_class = StorageCartFilter


class KittingPlanViewSet(viewsets.ModelViewSet):
	queryset = KittingPlan.objects.all()
	serializer_class = KittingPlanSerializer
	filter_class = KittingPlanFilter


class OrderItemViewSet(viewsets.ModelViewSet):
	queryset = OrderItem.objects.all()
	serializer_class = OrderItemSerializer
	filter_class = OrderItemFilter


class ApproveViewSet(viewsets.ModelViewSet):
	queryset = Approve.objects.all()
	serializer_class = ApproveSerializer
	filter_class = ApproveFilter


class OrderInfoViewSet(viewsets.ModelViewSet):
	queryset = OrderInfo.objects.all()
	serializer_class = OrderInfoSerializer
	filter_class = OrderInfoFilter


@login_required()
def top_page(request):
	return render(request, 'top.html')


class StorageItemListView(LoginRequiredMixin, ListView):
	model = StorageItem
	template_name = 'stock/storage_list.html'
	paginate_by = 30
	ordering = 'order_number'

	def get_context_data(self, *, object_list=None, **kwargs):
		context = super().get_context_data(**kwargs)
		base = Base.objects.all()
		context['base_list'] = base
		return context


class StorageItemDetailView(LoginRequiredMixin, DetailView):
	model = StorageItem
	template_name = 'stock/storage_detail.html'


class StorageItemCreateView(LoginRequiredMixin, BSModalCreateView):
	model = StorageItem
	template_name = 'snippets/create_modal.html'
	form_class = StorageItemBSModalForm
	success_url = reverse_lazy('stock:storage_list')


class StorageItemUpdateView(LoginRequiredMixin, BSModalUpdateView):
	model = StorageItem
	template_name = 'snippets/update_modal.html'
	form_class = StorageItemUpdateBSModalForm
	success_url = reverse_lazy('stock:storage_list')


class OptionCreateView(LoginRequiredMixin, BSModalCreateView):
	model = Option
	template_name = 'snippets/create_modal.html'
	form_class = OptionCreateBSModalForm
	success_url = reverse_lazy('stock:storage_list')


class StorageCartView(LoginRequiredMixin, TemplateView):
	model = StorageCart
	template_name = 'stock/cart.html'

	def get_context_data(self, **kwargs):
		context = super(StorageCartView, self).get_context_data(**kwargs)
		storage_cart = StorageCart.objects.get_or_none(requester__user__pk=self.request.user.pk, ordered=False)
		context['storagecart'] = storage_cart
		if storage_cart is not None:
			context['count'] = len(storage_cart.order_item.all())
		else:
			context['count'] = 0
		return context


def get_obj(resource, requester, pk):
	"""

	:param resource:
	:param requester:
	:param pk:
	:return:
	"""
	# cart = StorageCart.objects.filter(requester=requester, ordered=False)
	cart = StorageCart.objects.get_or_none(requester=requester, ordered=False)
	if resource == 'add':
		kitting_plan = KittingPlan.objects.get_or_none(name='標準')
		item = get_object_or_404(StorageItem, pk=pk)
		order_item, created = OrderItem.objects.get_or_create(
			storage_item=item,
			kitting_plan=kitting_plan,
			ordered=False,
			requester=requester
		)
	else:
		order_item = get_object_or_404(OrderItem, pk=pk)
	obj_data = {
		# 'cart': cart,
		'cart': cart,
		'order_item': order_item
	}
	return obj_data


@login_required()
def add_item(request, pk):
	"""
	カートにアイテムを追加する
	:param request:
	:param pk:
	:return:
	"""
	requester = Requester.objects.get(user=request.user)
	obj_data = get_obj('add', requester, pk)
	order_item = obj_data['order_item']
	cart = obj_data['cart']
	storage_quantity = order_item.storage_item.quantity
	if storage_quantity > 0:
		order_item.storage_item.quantity -= 1
		order_item.storage_item.save()
		for option in order_item.storage_item.option.all():
			option.quantity -= 1
			option.save()
	else:
		return redirect('stock:cart')
	if cart is not None:
		storage_cart = cart
		if storage_cart.order_item.filter(storage_item__pk=order_item.storage_item.pk).exists():
			order_item.quantity += 1
			order_item.save()
		else:
			storage_cart.order_item.add(order_item)
	else:
		storage_cart = StorageCart.objects.create(
			requester=requester
		)
		storage_cart.order_item.add(order_item)
	storage_cart.save()
	return redirect('stock:cart')


@login_required()
def reduce_cart(request, pk):
	"""
	カートからアイテムを減らす
	:param request:
	:param pk:
	:return:
	"""
	requester = Requester.objects.get(user=request.user)
	obj_data = get_obj('reduce', requester, pk)
	order_item = obj_data['order_item']
	cart = obj_data['cart']
	if cart is not None:
		order_item.storage_item.quantity += 1
		order_item.storage_item.save()
		for option in order_item.storage_item.option.all():
			option.quantity += 1
			option.save()
		if order_item.quantity > 1:
			order_item.quantity -= 1
			order_item.save()
		else:
			order_item.delete()
	else:
		pass
	return redirect('stock:cart')


@login_required()
def remove_cart(request, pk):
	"""
	カートからアイテムを削除する
	:param request:
	:param pk:
	:return:
	"""
	requester = Requester.objects.get(user=request.user)
	obj_data = get_obj('remove', requester, pk)
	order_item = obj_data['order_item']
	cart = obj_data['cart']
	if cart is not None:
		quantity = order_item.quantity
		order_item.storage_item.quantity += quantity
		order_item.storage_item.save()
		for option in order_item.storage_item.option.all():
			option.quantity += quantity
			option.save()
		cart_items = cart.order_item.all().count()
		if cart_items <= 1:
			order_item.delete()
			cart.delete()
		else:
			order_item.delete()
		cart.save()
	else:
		pass
	return redirect('stock:cart')


class ApproveView(LoginRequiredMixin, TemplateView):

	def get(self, request, *args, **kwargs):
		requester = Requester.objects.get(user=request.user)
		api_url = 'http://127.0.0.1:8000/api/v1/stock/orderItem/'
		cart = StorageCart.objects.get(requester=requester, ordered=False)
		cart.save()
		approve = Approve.objects.filter(requester=requester).last()
		kitting_plan = KittingPlan.objects.all()
		context = {
			'cart': cart,
			'kitting_plan': kitting_plan,
			'api_url': api_url
		}
		if approve is not None:
			context['approve'] = approve
		else:
			context['approve'] = None
		return render(request, 'stock/approve.html', context)


class ApproveCreateView(LoginRequiredMixin, BSModalCreateView):
	model = Approve
	template_name = 'snippets/create_modal.html'
	form_class = ApproveBSModalForm
	success_url = reverse_lazy('stock:approve')


class ApproveUpdateView(LoginRequiredMixin, BSModalUpdateView):
	model = Approve
	template_name = 'snippets/update_modal.html'
	form_class = ApproveBSModalForm
	success_url = reverse_lazy('stock:approve')


@login_required()
def add_order_info(request):
	"""
	approve画面から確定でorder_infoに登録する
	:param request:
	:return:
	"""
	requester = Requester.objects.get(user=request.user)
	approve = Approve.objects.get(requester=requester)
	cart = StorageCart.objects.get(requester=requester, ordered=False)
	date = datetime.datetime.now()
	str_date = date.strftime('%Y%m%d')
	order_info_obj = OrderInfo.objects.all().last()
	if order_info_obj is not None:
		if str_date in order_info_obj.number:
			old_branch = order_info_obj.number.split('-')[1]
			order_branch = int(old_branch) + 1
		else:
			order_branch = 1
	else:
		order_branch = 1
	for order_item in cart.order_item.all():
		order_item.storage_item.quantity -= order_item.quantity
		order_item.storage_item.save()
		for option in order_item.storage_item.option.all():
			option.quantity -= order_item.quantity
			option.save()
	number = str_date + '-' + str(order_branch)
	order_info, update = OrderInfo.objects.update_or_create(
		number=number,
		requester=requester,
		approve=approve,
		storage_cart=cart,
	)
	for order_item in cart.order_item.all():
		order_item.ordered = True
		order_item.save()
	cart.ordered = True
	cart.save()
	return redirect('stock:confirm', order_info.pk)


class ConfirmView(LoginRequiredMixin, TemplateView):
	template_name = 'stock/confirm.html'

	def get_context_data(self, *args, **kwargs):
		context = super(ConfirmView, self).get_context_data(**kwargs)
		order_info = OrderInfo.objects.get(pk=kwargs['pk'])
		order_item = order_info.storage_cart.order_item.all()
		orderitem_list = {}
		# for order_item in order_info.storage_cart.order_item.all():
		# 	orderitem_list[order_item.due_at.strftime('%Y年%m月%d日')] = []
		# 	orderitem_list[order_item.due_at.strftime('%Y年%m月%d日')].append(order_item)
		for i in range(len(order_item)):
			due = order_item[i].due_at
			if i > 0:
				if due != order_item[i - 1].due_at:
					orderitem_list[due.strftime('%Y年%m月%d日')] = []
			else:
				orderitem_list[due.strftime('%Y年%m月%d日')] = []
			orderitem_list[due.strftime('%Y年%m月%d日')].append(order_item[i])
		context['orderitem_list'] = orderitem_list
		context['order_info'] = order_info
		return context


class MyOrderInfoView(LoginRequiredMixin, ListView):
	model = OrderInfo
	template_name = 'stock/orderinfo_mypage.html'

	def get_queryset(self, *args, **kwargs):
		requester = Requester.objects.get(user=self.request.user)
		queryset = super(MyOrderInfoView, self).get_queryset()
		qs = queryset.filter(requester=requester).order_by('-pk')
		return qs


class OrderInfoDetailView(LoginRequiredMixin, DetailView):
	model = OrderInfo
	template_name = 'stock/orderinfo_detail.html'


class OrderInfoSelectView(LoginRequiredMixin, DetailView):
	model = OrderInfo
	template_name = 'stock/orderinfo_delete_select.html'

	@staticmethod
	def post(request, *args, **kwargs):
		order_item_pks = request.POST.getlist('selectItem')
		context = {
			'order_item_list': []
		}
		for pk in order_item_pks:
			order_item = OrderItem.objects.get(pk=pk)
			context['order_item_list'].append(order_item)
			order_item.storage_item.quantity += order_item.quantity
			order_item.storage_item.save()
			order_item.delete()
		order_info = OrderInfo.objects.get(pk=kwargs['pk'])
		order_item = order_info.storage_cart.order_item.all()
		if len(order_item) == 0:
			order_info.storage_cart.delete()
			order_info.delete()
		return render(request, 'stock/orderinfo_delete.html', context)


class ChangeQuantity(LoginRequiredMixin, TemplateView):
	def get(self, request, **kwargs):
		order_item = OrderItem.objects.get(pk=kwargs['pk'])
		quantity = order_item.quantity
		storage_quantity = order_item.storage_item.quantity
		context = {
			'storage_quantity': storage_quantity,
			'quantity': quantity
		}
		return render(request, 'snippets/quantity_modal.html', context)

	@staticmethod
	def post(request, **kwargs):
		order_item = OrderItem.objects.get(pk=kwargs['pk'])
		storage_item = order_item.storage_item
		post_quantity = int(request.POST['quantity'])
		after_quantity = order_item.quantity - post_quantity
		storage_item.quantity += after_quantity
		storage_item.save()
		order_item.quantity = post_quantity
		order_item.save()
		order_info = OrderInfo.objects.get(storage_cart__order_item=order_item)
		return redirect('stock:orderinfo_detail', order_info.pk)


def create_storage_data(request):
	"""
	データ取り込み用
	:param request:
	:return:
	"""
	file = r'D:\Users\19020081\Documents\【貯蔵品】PC在庫リスト.xlsx'
	wb = px.load_workbook(file)
	ws = wb.worksheets[0]
	# max_row = ws.max_row
	max_row = 442
	for i in range(3, max_row):
		b_value = ws['B' + str(i)].value
		active = ws['C' + str(i)].value
		base = ws['E' + str(i)].value
		order_number = ws['F' + str(i)].value
		date = ws['G' + str(i)].value
		item_name = ws['H' + str(i)].value
		model_number = ws['I' + str(i)].value
		price = int(ws['K' + str(i)].value)
		remarks = ws['L' + str(i)].value
		# delivery_date
		delivery_date = date.date().isoformat()
		# category, numpad
		category_numpad_size = get_category_and_numpad_and_size(b_value)
		category = category_numpad_size['category']
		size = category_numpad_size['size']
		# maker, name
		split_name = item_name.split(' ')
		maker = ''
		name = ''
		for n in range(len(split_name)):
			if n == 0:
				maker = split_name[n]
			else:
				name += split_name[n]
		if active == '空':
			pc = PC.objects.get_or_create(
				category=category,
				maker=maker,
				name=name,
				model_number=model_number
			)
			pc_detail = PCDetail.objects.get_or_create(
				pc=pc[0],
				cpu_id=1,
				storage_id=2,
				size=size,
			)
			get_base = Base.objects.get_or_create(
				name=base
			)

			stock_storage, create_storage = StorageItem.objects.get_or_create(
				order_number=order_number,
				item=pc_detail[0],
				price=price,
				base=get_base[0],
				registration_at=delivery_date,
			)
			if create_storage:
				stock_storage.quantity = 1
			else:
				stock_storage.quantity += 1
			stock_storage.remarks = remarks
			stock_storage.save()
	return redirect('stock:storage_list')


def get_category_and_numpad_and_size(value):
	"""
	categoryを分別する
	:param value:
	:return category_and_numpad_and_size:
	"""
	category_and_numpad_and_size = {
		'size': None
	}
	if 'ノート' in value:
		category_and_numpad_and_size['category'] = '1'
		if 'B5' in value:
			category_and_numpad_and_size['numpad'] = False
			category_and_numpad_and_size['size'] = 13
		elif 'A4' in value:
			category_and_numpad_and_size['numpad'] = True
			category_and_numpad_and_size['size'] = 15
	elif '小型' in value:
		category_and_numpad_and_size['category'] = '3'
		category_and_numpad_and_size['numpad'] = False
	return category_and_numpad_and_size
