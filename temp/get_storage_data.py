# -*- coding:utf-8 -*-
import openpyxl as px
import datetime
import os
import glob
import requests
import json


def search_file(m):
	"""

	:param m:
	:return file:
	"""
	file_name = os.environ['USERPROFILE'] + r'\Downloads\【貯蔵品】PC在庫リスト.xlsx'
	return file_name


def post_storage_data():
	"""
	エクセルファイルからデータを取得
	:param :
	:return:
	"""
	file = os.environ['USERPROFILE'] + r'\Downloads\【貯蔵品】PC在庫リスト.xlsx'
	device_url = 'http://127.0.0.1:8000/api/v1/device/item/'
	storage_url = 'http://127.0.0.1:8000/api/v1/stock/storage/'
	wb = px.load_workbook(file)
	ws = wb.worksheets[0]
	max_row = ws.max_row + 1

	for i in range(3, max_row):
		b_value = ws['B' + str(i)].value
		active = ws['C' + str(i)].value
		base = ws['E' + str(i)].value
		order_number = ws['F' + str(i)].value
		date = ws['G' + str(i)].value
		item_name = ws['H' + str(i)].value
		model_number = ws['I' + str(i)].value
		price = int(ws['K' + str(i)].value)
		remarks = ws['L' + str(i)].value
		# delivery_date
		delivery_date = date.date().isoformat()
		# category, numpad
		category_numpad_size = get_category_and_numpad_and_size(b_value)
		category = category_numpad_size['category']
		numpad = category_numpad_size['numpad']
		size = category_numpad_size['size']
		# maker, name
		split_name = item_name.split(' ')
		maker = ''
		name = ''
		for n in range(len(split_name)):
			if n == 0:
				maker = split_name[n]
			else:
				name += split_name[n]
		if active == '空':
			headers = {'content-type': 'application/json'}
			storage_results = get_api_date(storage_url, headers)
			device_results = get_api_date(device_url, headers)
			post_pc = {
				'category': category,
				'maker': maker,
				'name': name,
				'model_number': model_number
			}

			post_device = {
				'pc': post_pc,
				'spec': {
					'cpu': {
						'maker': 'intel',
						'name': 'core i5',
						'gen': 8
					},
					'memory': '1',
					'storage': {
						'type': 'SSD',
						'size': 256
					},
					'size': size,
					'camera': False,
					'fingerprint': False,
					'numpad': numpad,
					'lan': False,
					'usb': None,
					'hdmi': None,
					'vga': None
				},
				'remarks': ''
			}
			for storage in storage_results:
				post_storage = {
					'order_number': order_number,
					'item': post_device,
					'price': price,
					'quantity': '',
					'option': [],
					'base': {
						'name': base
					},
					'delivery_date': delivery_date,
					'remarks': ''
				}
				if storage['item']['pc'] != post_pc:
					post_device_api(post_device, device_results, device_url, headers)
					post_storage['quantity'] = 1
					storage_res = requests.post(storage_url, data=json.dumps(post_storage), headers=headers)
					print(sotrage_res.status_code)
				else:
					post_storage['quantity'] = storage['quantity'] + 1
					storage_res = requests.post(storage_url, data=json.dumps(post_storage), headers=headers)
					print(storage_res.status_code)


def get_category_and_numpad_and_size(value):
	"""
	categoryを分別する
	:param value:
	:return:
	"""
	category_and_numpad_and_size = {}
	if 'ノート' in value:
		category_and_numpad_and_size['category'] = '1'
		if 'B5' in value:
			category_and_numpad_and_size['numpad'] = False
			category_and_numpad_and_size['size'] = 13
		elif 'A4' in value:
			category_and_numpad_and_size['numpad'] = True
			category_and_numpad_and_size['size'] = 15
	elif '小型' in value:
		category_and_numpad_and_size['category'] = '3'
		category_and_numpad_and_size['numpad'] = False
	return category_and_numpad_and_size


def get_api_date(url, headers):
	"""
a	apiからデータを持ってくる
	:param url:
	:param headers:
	:return:
	"""
	# headers = {'content-type': 'application/json'}
	response = requests.get(url, headers=headers)
	res_text = json.loads(response.text)
	results = res_text['results']
	return results


def post_device_api(data, result, url, headers):
	"""
	deviceを確認してなければpost
	:param data:
	:param result:
	:param url:
	:param headers:
	:return:
	"""
	for device in result:
		if data != device:
			res = requests.post(url, data=json.dumps(data), headers=headers)
			print(res.text)


if __name__ == '__main__':
	post_storage_data()
